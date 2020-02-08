import openpyxl
import paramiko
import time
import getpass
import os
from config_file import host_conf #it will import the list of commands that you want to apply from config_file.py

loc = "IPs.xlsx" #your excel file path that contain your devices IPs
wb = openpyxl.load_workbook(loc) 
sheet = wb["sheet1"] #your sheet name

for i in  range(sheet.max_row):
	ip=sheet.cell(i+1, 1)
	print (ip.value)
	twrssh = paramiko.SSHClient()
	twrssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
	twrssh.connect(ip.value, port=22, username='cisco', password='cisco')
	remote = twrssh.invoke_shell()
	time.sleep(4)
	for command in host_conf:
		nip=sheet.cell(i+1, 2)
		remote.send(' %s \n' %command)
		time.sleep(4)
		buf = remote.recv(65000)
		outp = buf.decode('utf-8')
		nip=outp
		f = open('output.txt', 'a') #it will create text file contain output 
		f.write(outp)
		f.close()
	twrssh.close()
